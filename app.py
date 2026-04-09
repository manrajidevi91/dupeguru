import os
import json
import logging
import shutil
from flask import Flask, render_template, request, jsonify, send_file
from pathlib import Path
from send2trash import send2trash

# dupeGuru core imports
from core import fs, scanner, engine
from core.pe.scanner import ScannerPE
from core.pe.photo import Photo as PhotoFile
from core.me.scanner import ScannerME
from core.me.fs import MusicFile
from core.se.scanner import ScannerSE
from core.se.fs import File as StandardFile

app = Flask(__name__)

@app.route('/api/browse_os', methods=['GET'])
def browse_os_directory():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        folder_path = filedialog.askdirectory(parent=root, title="Select Folder for dupeGuru")
        root.destroy()
        
        if folder_path:
            return jsonify({"status": "success", "path": os.path.abspath(folder_path)})
        return jsonify({"status": "cancel"})
    except Exception as e:
        logger.error(f"Error opening native browse: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global state for the scan session
class ScanSession:
    def __init__(self):
        self.folders = []
        self.results = [] # List of groups
        self.scan_type = 10 # Default to FuzzyBlock
        self.threshold = 95
        self.scanning = False
        self.appdata = Path(os.environ.get('APPDATA', '.')) / 'dupeguru-web'
        self.appdata.mkdir(parents=True, exist_ok=True)
        self.cache_path = str(self.appdata / 'cached_pictures.db')

session = ScanSession()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/folders', methods=['GET', 'POST', 'DELETE'])
def manage_folders():
    if request.method == 'POST':
        data = request.json
        folder_path = data.get('path')
        if folder_path and os.path.isdir(folder_path):
            if folder_path not in session.folders:
                session.folders.append(folder_path)
            return jsonify({"status": "success", "folders": session.folders})
        return jsonify({"status": "error", "message": "Invalid folder path"}), 400
    
    elif request.method == 'DELETE':
        data = request.json
        folder_path = data.get('path')
        if folder_path in session.folders:
            session.folders.remove(folder_path)
        return jsonify({"status": "success", "folders": session.folders})
    
    return jsonify({"folders": session.folders})

@app.route('/api/scan', methods=['POST'])
def run_scan():
    data = request.json or {}
    mode = data.get('mode', 'picture')
    algorithm = int(data.get('algorithm', 10))
    threshold = int(data.get('threshold', 95))

    if not session.folders:
        return jsonify({"status": "error", "message": "No folders selected"}), 400
    
    session.scanning = True
    try:
        # 1. Select Engine and File Class
        if mode == 'music':
            scanner_cls = ScannerME
            file_classes = [MusicFile]
        elif mode == 'standard':
            scanner_cls = ScannerSE
            file_classes = [StandardFile]
        else:
            scanner_cls = ScannerPE
            file_classes = [PhotoFile]

        # 2. Collect files from folders
        all_files = []
        for folder in session.folders:
            folder_path = Path(folder)
            for root, dirs, files in os.walk(folder_path):
                for name in files:
                    file_path = Path(root) / name
                    try:
                        f = fs.get_file(file_path, file_classes)
                        if f:
                            all_files.append(f)
                    except Exception as e:
                        logger.error(f"Error reading {file_path}: {e}")

        # 3. Run Scanner
        s = scanner_cls()
        s.scan_type = algorithm
        s.min_match_percentage = threshold
        if mode == 'picture':
            s.cache_path = session.cache_path
        
        groups = s.get_dupe_groups(all_files)
        
        # 4. Format results for JSON
        session.results = []
        for i, group in enumerate(groups):
            group_data = {
                "id": i + 1,
                "ref": format_file(group.ref),
                "duplicates": [format_file(dupe, group) for dupe in group.dupes]
            }
            session.results.append(group_data)
            
        session.scanning = False
        return jsonify({"status": "success", "groups": session.results})
    
    except Exception as e:
        session.scanning = False
        logger.exception("Scan failed")
        return jsonify({"status": "error", "message": str(e)}), 500

def format_file(f, group=None):
    # Ensure metadata is read
    f._read_all_info()
    
    data = {
        "path": str(f.path),
        "name": f.name,
        "size": f.size,
        "folder": str(f.path.parent),
    }
    if hasattr(f, 'dimensions'):
        data["dimensions"] = f"{f.dimensions[0]} x {f.dimensions[1]}"
    
    if group:
        match = group.get_match_of(f)
        if match:
            data["percentage"] = match.percentage
            
    return data

@app.route('/api/image')
def get_image():
    path = request.args.get('path')
    if path and os.path.exists(path):
        return send_file(path)
    return "Not Found", 404

@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        groups = data.get('groups', [])
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Duplicate Report"
        
        # Color Palette (Modern Pastels)
        palette = [
            "EBF2FF", # Soft Blue
            "F0FFF4", # Hint of Mint
            "FFF5F5", # Soft Rose
            "FAF5FF", # Lavender
            "FFFFF0", # Ivory
            "F0FBFF", # Sky
            "FFF9E6", # Champagne
        ]
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Headers
        headers = ["Group ID", "Retained Master File", "Deleted Duplicate File"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25

        for i, group in enumerate(groups):
            group_id = group.get('group_id')
            # Extract only filename
            master = os.path.basename(group.get('master_path', ''))
            dupes = group.get('duplicates', [])
            
            if not dupes:
                continue
            
            # Pick a color for this group
            fill_color = palette[i % len(palette)]
            group_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            content_font = Font(size=10, color="1E293B")
            
            for dupe_path in dupes:
                # Extract only filename
                dupe_name = os.path.basename(dupe_path)
                
                ws.append([group_id, master, dupe_name])
                
                # Apply styling to the newly added row
                for cell in ws[ws.max_row]:
                    cell.fill = group_fill
                    cell.font = content_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", indent=1)

        # Auto-adjust columns width
        column_widths = [10, 40, 40] # Default minimums
        for i, col in enumerate(ws.columns):
            max_length = column_widths[i]
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 4, 80)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='dupeGuru_Audit_Report.xlsx'
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/delete', methods=['POST'])
def delete_files():
    data = request.json
    file_paths = data.get('paths', [])
    success_count = 0
    errors = []
    
    for path in file_paths:
        try:
            if os.path.exists(path):
                send2trash(path)
                success_count += 1
        except Exception as e:
            errors.append({"path": path, "error": str(e)})
            
    return jsonify({"status": "success", "deleted": success_count, "errors": errors})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5010, debug=True, use_reloader=False)
